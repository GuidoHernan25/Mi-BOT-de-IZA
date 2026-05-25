from __future__ import annotations

from collections import OrderedDict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\guido\Downloads\Copia de Operacion Logistica.xlsx")
OUTPUT_DIR = ROOT / "outputs"
OUTPUT = OUTPUT_DIR / "Operacion Logistica Mejorada.xlsx"


DARK = "16324F"
TEAL = "0F766E"
LIGHT = "EAF2F8"
GRAY = "F3F4F6"
OK = "BBF7D0"
WARN = "FDE68A"
BAD = "FCA5A5"
WHITE = "FFFFFF"


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def excel_serial_to_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=float(value))
    return value or ""


def style_header(ws, row=1, max_col=None):
    max_col = max_col or ws.max_column
    fill = PatternFill("solid", fgColor=TEAL)
    font = Font(color=WHITE, bold=True)
    for cell in ws[row][:max_col]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24


def add_table(ws, name: str, ref: str):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_list_validation(ws, range_ref: str, formula: str):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(range_ref)


def make_workbook():
    source = load_workbook(SOURCE, data_only=False)
    inv_ws = source["INVENTARIO"]
    salidas_ws = source["SALIDAS"]
    devol_ws = source["DEVOLUCIONES"]

    inventory_rows = []
    for code, tipo, desc, estado, *_ in inv_ws.iter_rows(min_row=2, values_only=True):
        code = clean(code)
        if not code:
            continue
        inventory_rows.append(
            {
                "code": code,
                "type": clean(tipo),
                "description": clean(desc),
                "state": clean(estado) or "Disponible",
            }
        )

    products = OrderedDict()
    for row in inventory_rows:
        tipo = row["type"] or "SIN-TIPO"
        if tipo not in products:
            products[tipo] = {"type": tipo, "description": row["description"] or tipo, "total": 0}
        products[tipo]["total"] += 1
        if not products[tipo]["description"] and row["description"]:
            products[tipo]["description"] = row["description"]

    product_rows = sorted(products.values(), key=lambda item: item["type"])

    salida_rows = []
    for fecha, evento, code, *_ in salidas_ws.iter_rows(min_row=2, values_only=True):
        code = clean(code)
        if code:
            salida_rows.append([excel_serial_to_datetime(fecha), clean(evento), "", code])

    devol_rows = []
    for fecha, evento, code, *_ in devol_ws.iter_rows(min_row=2, values_only=True):
        code = clean(code)
        if code:
            devol_rows.append([excel_serial_to_datetime(fecha), clean(evento), "", code])

    events = sorted({row[1] for row in salida_rows + devol_rows if row[1]}) or ["Evento demo"]

    wb = Workbook()
    wb.remove(wb.active)
    sheets = {
        name: wb.create_sheet(name)
        for name in [
            "INICIO",
            "PRODUCTOS",
            "INVENTARIO",
            "PEDIDOS",
            "PEDIDO_ITEMS",
            "SALIDAS",
            "DEVOLUCIONES",
            "CONTROL_EVENTOS",
            "STOCK EN VIVO",
            "REMITO_DIGITAL",
            "HISTORIAL_REMITOS",
            "LISTAS",
        ]
    }

    for ws in sheets.values():
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

    inicio = sheets["INICIO"]
    inicio.merge_cells("A1:H1")
    inicio["A1"] = "OPERACION LOGISTICA - PANEL SIMPLE"
    inicio["A1"].fill = PatternFill("solid", fgColor=DARK)
    inicio["A1"].font = Font(color=WHITE, bold=True, size=18)
    inicio["A1"].alignment = Alignment(horizontal="center")
    inicio.merge_cells("A3:H3")
    inicio["A3"] = "Resumen rapido"
    inicio["A3"].fill = PatternFill("solid", fgColor=LIGHT)
    inicio["A3"].font = Font(color=DARK, bold=True, size=13)
    for row in [
        ("Codigos inventario", len(inventory_rows)),
        ("Tipos de producto", len(product_rows)),
        ("Eventos activos", len(events)),
        ("Salidas cargadas", len(salida_rows)),
        ("Devoluciones cargadas", len(devol_rows)),
    ]:
        inicio.append(row)
    inicio["D5"] = "Hojas de uso diario"
    inicio["E5"] = "PEDIDOS"
    inicio["F5"] = "SALIDAS"
    inicio["G5"] = "DEVOLUCIONES"
    inicio["H5"] = "STOCK EN VIVO"
    inicio["D7"] = "Alertas automaticas"
    inicio["E7"] = "CODIGO NO EXISTE"
    inicio["F7"] = "DUPLICADO"
    inicio["G7"] = "NO PEDIDO"
    inicio["H7"] = "EXCEDE PEDIDO"
    for cell in inicio["A"]:
        if cell.row >= 5 and cell.value:
            cell.fill = PatternFill("solid", fgColor=GRAY)
            cell.font = Font(bold=True)
    set_widths(inicio, [22, 14, 4, 22, 20, 20, 20, 20])

    productos = sheets["PRODUCTOS"]
    productos.append(["Tipo", "Descripcion", "Total inventario", "Activo"])
    for item in product_rows:
        productos.append([item["type"], item["description"], item["total"], "SI"])
    style_header(productos)
    add_table(productos, "TablaProductos", f"A1:D{productos.max_row}")
    set_widths(productos, [22, 34, 16, 10])

    inventario = sheets["INVENTARIO"]
    inventario.append(["Codigo", "Tipo", "Descripcion", "Estado", "Observaciones"])
    for row in inventory_rows:
        inventario.append([row["code"], row["type"], row["description"], row["state"], ""])
    style_header(inventario)
    add_table(inventario, "TablaInventario", f"A1:E{inventario.max_row}")
    set_widths(inventario, [20, 20, 36, 16, 30])
    add_list_validation(inventario, f"D2:D{inventario.max_row}", '"Disponible,Reservado,Afuera,Mantenimiento,Perdido"')

    listas = sheets["LISTAS"]
    listas.append(["Tipos", "Eventos", "Estados Pedido"])
    for idx in range(max(len(product_rows), len(events), 7)):
        listas.append(
            [
                product_rows[idx]["type"] if idx < len(product_rows) else "",
                events[idx] if idx < len(events) else "",
                ["Borrador", "Confirmado", "En preparacion", "Despachado", "Devuelto", "Cerrado", "Cancelado"][idx]
                if idx < 7
                else "",
            ]
        )
    style_header(listas)
    set_widths(listas, [24, 28, 20])
    listas.sheet_state = "hidden"

    pedidos = sheets["PEDIDOS"]
    pedidos.append(["Pedido", "Fecha", "Evento", "Solicitante", "Estado", "Notas", "Validacion"])
    for idx, event in enumerate(events, start=1):
        pedidos.append([f"PED-{idx:04d}", datetime.today(), event, "", "Confirmado" if idx == 1 else "Borrador", "", ""])
    for row in range(pedidos.max_row + 1, 502):
        pedidos.append(["", "", "", "", "", "", ""])
    style_header(pedidos)
    add_table(pedidos, "TablaPedidos", "A1:G501")
    set_widths(pedidos, [16, 13, 24, 22, 18, 36, 18])
    add_list_validation(pedidos, "C2:C501", f"LISTAS!$B$2:$B${len(events)+1}")
    add_list_validation(pedidos, "E2:E501", "LISTAS!$C$2:$C$8")
    for cell in pedidos["B"]:
        cell.number_format = "yyyy-mm-dd"

    pedido_items = sheets["PEDIDO_ITEMS"]
    pedido_items.append(["Pedido", "Tipo", "Cantidad pedida", "Cantidad salida", "Faltan", "Estado"])
    for row in range(2, 502):
        pedido_items.append(
            [
                "",
                "",
                "",
                f'=IF(A{row}="","",COUNTIFS(SALIDAS!$C$2:$C$1000,A{row},SALIDAS!$E$2:$E$1000,B{row},SALIDAS!$G$2:$G$1000,"OK*"))',
                f'=IF(A{row}="","",MAX(0,C{row}-D{row}))',
                f'=IF(A{row}="","",IF(E{row}=0,"COMPLETO","FALTA"))',
            ]
        )
    style_header(pedido_items)
    add_table(pedido_items, "TablaPedidoItems", "A1:F501")
    set_widths(pedido_items, [16, 22, 18, 18, 12, 16])
    add_list_validation(pedido_items, "B2:B501", f"LISTAS!$A$2:$A${len(product_rows)+1}")

    salidas = sheets["SALIDAS"]
    salidas.append(["Fecha", "Evento", "Pedido", "Codigo", "Tipo detectado", "Descripcion", "Validacion"])
    for idx in range(999):
        base = salida_rows[idx] if idx < len(salida_rows) else ["", "", "", ""]
        row = idx + 2
        salidas.append(
            base
            + [
                f'=IF(D{row}="","",IFERROR(XLOOKUP(TRIM(D{row}),INVENTARIO!$A$2:$A$2000,INVENTARIO!$B$2:$B$2000),"NO ENCONTRADO"))',
                f'=IF(D{row}="","",IFERROR(XLOOKUP(TRIM(D{row}),INVENTARIO!$A$2:$A$2000,INVENTARIO!$C$2:$C$2000),""))',
                f'=IF(D{row}="","",IF(E{row}="NO ENCONTRADO","CODIGO NO EXISTE",IF(COUNTIF($D$2:D{row},D{row})>1,"DUPLICADO",IF(C{row}="","OK SIN PEDIDO",IF(COUNTIFS(PEDIDO_ITEMS!$A$2:$A$501,C{row},PEDIDO_ITEMS!$B$2:$B$501,E{row})=0,"NO PEDIDO",IF(COUNTIFS($C$2:C{row},C{row},$E$2:E{row},E{row})>SUMIFS(PEDIDO_ITEMS!$C$2:$C$501,PEDIDO_ITEMS!$A$2:$A$501,C{row},PEDIDO_ITEMS!$B$2:$B$501,E{row}),"EXCEDE PEDIDO","OK"))))))',
            ]
        )
    style_header(salidas)
    add_table(salidas, "TablaSalidas", "A1:G1000")
    set_widths(salidas, [18, 24, 16, 20, 20, 36, 22])
    add_list_validation(salidas, "B2:B1000", f"LISTAS!$B$2:$B${len(events)+1}")
    for cell in salidas["A"]:
        cell.number_format = "yyyy-mm-dd hh:mm"

    devoluciones = sheets["DEVOLUCIONES"]
    devoluciones.append(["Fecha", "Evento", "Pedido", "Codigo", "Tipo detectado", "Validacion"])
    for idx in range(999):
        base = devol_rows[idx] if idx < len(devol_rows) else ["", "", "", ""]
        row = idx + 2
        devoluciones.append(
            base
            + [
                f'=IF(D{row}="","",IFERROR(XLOOKUP(TRIM(D{row}),INVENTARIO!$A$2:$A$2000,INVENTARIO!$B$2:$B$2000),"NO ENCONTRADO"))',
                f'=IF(D{row}="","",IF(E{row}="NO ENCONTRADO","CODIGO NO EXISTE",IF(COUNTIF($D$2:D{row},D{row})>1,"DEVOLUCION DUPLICADA",IF(COUNTIFS(SALIDAS!$B$2:$B$1000,B{row},SALIDAS!$D$2:$D$1000,D{row},SALIDAS!$G$2:$G$1000,"OK*")=0,"NO FIGURA EN SALIDA","OK"))))',
            ]
        )
    style_header(devoluciones)
    add_table(devoluciones, "TablaDevoluciones", "A1:F1000")
    set_widths(devoluciones, [18, 24, 16, 20, 20, 24])
    add_list_validation(devoluciones, "B2:B1000", f"LISTAS!$B$2:$B${len(events)+1}")
    for cell in devoluciones["A"]:
        cell.number_format = "yyyy-mm-dd hh:mm"

    control = sheets["CONTROL_EVENTOS"]
    control.append(["Evento", "Salieron OK", "Volvieron OK", "Faltan", "Codigos con alerta", "Estado"])
    for idx, event in enumerate(events, start=2):
        control.append(
            [
                event,
                f'=IF(A{idx}="","",COUNTIFS(SALIDAS!$B$2:$B$1000,A{idx},SALIDAS!$G$2:$G$1000,"OK*"))',
                f'=IF(A{idx}="","",COUNTIFS(DEVOLUCIONES!$B$2:$B$1000,A{idx},DEVOLUCIONES!$F$2:$F$1000,"OK"))',
                f'=IF(A{idx}="","",MAX(0,B{idx}-C{idx}))',
                f'=IF(A{idx}="","",COUNTIFS(SALIDAS!$B$2:$B$1000,A{idx},SALIDAS!$G$2:$G$1000,"<>OK",SALIDAS!$G$2:$G$1000,"<>OK SIN PEDIDO"))',
                f'=IF(A{idx}="","",IF(E{idx}>0,"REVISAR",IF(D{idx}=0,"COMPLETO","INCOMPLETO")))',
            ]
        )
    for idx in range(control.max_row + 1, 101):
        control.append(["", "", "", "", "", ""])
    style_header(control)
    add_table(control, "TablaControlEventos", "A1:F100")
    set_widths(control, [24, 16, 16, 12, 18, 18])

    stock = sheets["STOCK EN VIVO"]
    stock.append(["Tipo", "Total", "Reservado pedido", "Afuera", "En deposito", "Alerta"])
    for idx, item in enumerate(product_rows, start=2):
        stock.append(
            [
                item["type"],
                f'=IF(A{idx}="","",COUNTIF(INVENTARIO!$B$2:$B$2000,A{idx}))',
                f'=IF(A{idx}="","",MAX(0,SUMIFS(PEDIDO_ITEMS!$C$2:$C$501,PEDIDO_ITEMS!$B$2:$B$501,A{idx})-COUNTIFS(SALIDAS!$E$2:$E$1000,A{idx},SALIDAS!$G$2:$G$1000,"OK*")))',
                f'=IF(A{idx}="","",MAX(0,COUNTIFS(SALIDAS!$E$2:$E$1000,A{idx},SALIDAS!$G$2:$G$1000,"OK*")-COUNTIFS(DEVOLUCIONES!$E$2:$E$1000,A{idx},DEVOLUCIONES!$F$2:$F$1000,"OK")))',
                f'=IF(A{idx}="","",B{idx}-C{idx}-D{idx})',
                f'=IF(A{idx}="","",IF(E{idx}<0,"REVISAR",IF(E{idx}=0,"SIN STOCK","OK")))',
            ]
        )
    style_header(stock)
    add_table(stock, "TablaStockEnVivo", f"A1:F{stock.max_row}")
    set_widths(stock, [24, 12, 18, 12, 14, 14])

    remito = sheets["REMITO_DIGITAL"]
    remito.merge_cells("A1:J1")
    remito["A1"] = "REMITO DIGITAL"
    remito["A1"].fill = PatternFill("solid", fgColor=DARK)
    remito["A1"].font = Font(color=WHITE, bold=True, size=18)
    remito["A1"].alignment = Alignment(horizontal="center")
    for row in [
        ("Empresa", "IZANERGIA"),
        ("CUIT", "20-5104893-3"),
        ("Direccion", "Pizurno 1186, Del Viso"),
        ("Evento", ""),
        ("Pedido", ""),
    ]:
        remito.append(row)
    remito["H3"] = "N Remito"
    remito["H4"] = "Fecha"
    remito["I4"] = datetime.today()
    remito["H5"] = "Estado"
    remito["I5"] = "Borrador"
    remito["A10"] = "Tipo"
    remito["B10"] = "Descripcion"
    remito["C10"] = "Cantidad"
    remito["D10"] = "Unidad"
    remito["E10"] = "Observaciones"
    style_header(remito, 10, 5)
    for _ in range(11, 36):
        remito.append(["", "", "", "unidad", ""])
    remito["B39"] = "Firma: ___________________________________________"
    set_widths(remito, [18, 34, 12, 12, 30, 8, 8, 16, 16, 16])

    historial = sheets["HISTORIAL_REMITOS"]
    historial.append(["N Remito", "Fecha", "Evento", "Pedido", "Tipo", "Descripcion", "Cantidad", "Unidad"])
    for _ in range(2, 502):
        historial.append(["", "", "", "", "", "", "", ""])
    style_header(historial)
    add_table(historial, "TablaHistorialRemitos", "A1:H501")
    set_widths(historial, [12, 14, 24, 16, 20, 36, 12, 10])

    thin = Side(style="thin", color="D1D5DB")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.border = Border(bottom=thin)
        ws.auto_filter.ref = ws.dimensions

    for ws_name, status_col in [
        ("SALIDAS", "G"),
        ("DEVOLUCIONES", "F"),
        ("CONTROL_EVENTOS", "F"),
        ("STOCK EN VIVO", "F"),
        ("PEDIDO_ITEMS", "F"),
    ]:
        ws = sheets[ws_name]
        max_row = ws.max_row
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{max_row}",
            FormulaRule(formula=[f'ISNUMBER(SEARCH("OK",{status_col}2))'], fill=PatternFill("solid", fgColor=OK)),
        )
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{max_row}",
            FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("NO ",{status_col}2)),ISNUMBER(SEARCH("REVISAR",{status_col}2)))'], fill=PatternFill("solid", fgColor=BAD)),
        )
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{max_row}",
            FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("FALTA",{status_col}2)),ISNUMBER(SEARCH("DUPLICADA",{status_col}2)),ISNUMBER(SEARCH("DUPLICADO",{status_col}2)))'], fill=PatternFill("solid", fgColor=WARN)),
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    output = make_workbook()
    print(output)
